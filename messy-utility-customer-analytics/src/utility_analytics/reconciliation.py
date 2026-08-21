"""Billing and meter reconciliation."""

from __future__ import annotations

from pathlib import Path

import polars as pl
from openpyxl import load_workbook


def load_tariffs(path: str | Path) -> pl.DataFrame:
    """Load tariff reference data from Excel."""

    workbook = load_workbook(Path(path), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    headers = rows[0]
    tariffs = pl.DataFrame([dict(zip(headers, row, strict=False)) for row in rows[1:]])
    return tariffs.with_columns(
        pl.col("customer_category").cast(pl.Utf8),
        pl.col("region").cast(pl.Utf8).str.to_uppercase(),
        pl.col("effective_from").cast(pl.Utf8),
        pl.col("effective_to").cast(pl.Utf8),
    )


def build_meter_history(meters: pl.DataFrame) -> pl.DataFrame:
    """Build effective-dated meter ownership history."""

    current = meters.select(
        pl.col("canonical_meter_id").alias("meter_id"),
        pl.col("canonical_customer_id").alias("customer_id"),
        pl.col("canonical_install_date").alias("valid_from"),
        pl.col("canonical_reassignment_effective_from").alias("valid_to"),
    )
    reassigned = meters.filter(pl.col("reassigned_to_canonical_customer_id").is_not_null()).select(
        pl.col("canonical_meter_id").alias("meter_id"),
        pl.col("reassigned_to_canonical_customer_id").alias("customer_id"),
        pl.col("canonical_reassignment_effective_from").alias("valid_from"),
        pl.lit(None).cast(pl.Utf8).alias("valid_to"),
    )
    return pl.concat([current, reassigned], how="vertical_relaxed")


def reconcile_bills(
    customers: pl.DataFrame,
    meters: pl.DataFrame,
    readings: pl.DataFrame,
    billing: pl.DataFrame,
    tariffs: pl.DataFrame,
) -> pl.DataFrame:
    """Reconcile billed amounts against meter-consumption and tariff expectations."""

    customer_meta = customers.select(
        "canonical_customer_id",
        pl.col("customer_category"),
        pl.col("normalized_location").alias("region"),
    )

    meter_history = build_meter_history(meters)
    reading_totals = readings.with_columns(pl.col("canonical_reading_date").str.slice(0, 7).alias("billing_period")).join(
        meter_history,
        left_on="canonical_meter_id",
        right_on="meter_id",
        how="left",
    ).with_columns(
        pl.coalesce([pl.col("customer_id"), pl.col("canonical_customer_id")]).alias("assigned_customer_id")
    ).group_by(["assigned_customer_id", "canonical_meter_id", "billing_period"]).agg(
        pl.col("consumption_kwh").sum().alias("meter_consumption_kwh")
    )

    reconciliation = billing.join(
        reading_totals,
        left_on=["canonical_customer_id", "canonical_meter_id", "canonical_billing_period"],
        right_on=["assigned_customer_id", "canonical_meter_id", "billing_period"],
        how="full",
    ).join(customer_meta, on="canonical_customer_id", how="left")

    reconciliation = reconciliation.join(tariffs, on=["customer_category", "region"], how="left")
    reconciliation = reconciliation.with_columns(
        pl.col("discount").fill_null(0),
        pl.when(pl.col("meter_consumption_kwh").is_not_null() & pl.col("energy_rate").is_not_null())
        .then((pl.col("meter_consumption_kwh") * pl.col("energy_rate") + pl.col("fixed_charge")) * (1 + pl.col("tax_rate")) - pl.col("discount"))
        .otherwise(None)
        .alias("expected_amount"),
        pl.col("bill_amount").alias("actual_amount"),
    )

    reconciliation = reconciliation.with_columns(
        pl.when(pl.col("bill_status") == "CANCELLED").then(pl.lit("CANCELLED"))
        .when(pl.col("bill_id").is_null()).then(pl.lit("MISSING_BILL"))
        .when(pl.col("meter_consumption_kwh").is_null()).then(pl.lit("NO_READING"))
        .when(pl.col("energy_rate").is_null()).then(pl.lit("INVALID_TARIFF"))
        .when(pl.col("actual_amount").is_null()).then(pl.lit("MISSING_BILL"))
        .when(
            (pl.col("expected_amount").is_not_null())
            & ((pl.col("actual_amount") - pl.col("expected_amount")).abs() / pl.col("expected_amount").abs().fill_null(1) <= 0.05)
        )
        .then(pl.lit("MATCH"))
        .otherwise(pl.lit("MISMATCH"))
        .alias("reconciliation_status"),
        (pl.col("actual_amount") - pl.col("expected_amount")).alias("difference"),
    ).with_columns(
        (pl.col("difference") / pl.col("expected_amount").abs().fill_null(1) * 100).alias("difference_pct")
    )

    return reconciliation
